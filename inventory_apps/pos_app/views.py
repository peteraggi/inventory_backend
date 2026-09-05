# pos_app/views.py

from rest_framework import generics, status, views, permissions, filters
from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes
from rest_framework.exceptions import ValidationError
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Sum, Count, Q, F, Avg
from django.db import transaction
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal
from drf_yasg.utils import swagger_auto_schema
from drf_yasg import openapi
import logging

from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import openpyxl

from .models import (
    Store,
    Role,
    Category,
    Product,
    Invoice,
    InvoiceItem,
    SyncLog,
    DailySales,
    Contact,
    Account,
    Journal,
    JournalEntry,
    JournalEntryLine,
    PurchaseOrder,
    PurchaseOrderLine,
    GoodsReceipt,
    GoodsReceiptLine,
    VendorBill,
    VendorBillLine,
    SalesOrder,
    SalesOrderLine,
    DeliveryNote,
    DeliveryNoteLine,
    SalesInvoice,
    SalesInvoiceLine,
    Payment,
    StockMove,
    StockLayer,
)
from .serializers import (
    StoreSerializer,
    StoreListSerializer,
    RoleSerializer,
    CategorySerializer,
    ProductSerializer,
    ProductListSerializer,
    InvoiceSerializer,
    InvoiceListSerializer,
    BulkInvoiceSyncSerializer,
    DashboardStatsSerializer,
    SalesReportSerializer,
    ProductReportSerializer,
    SyncLogSerializer,
    UserProfileSerializer,
    ContactSerializer,
    ContactListSerializer,
    AccountSerializer,
    JournalSerializer,
    JournalEntrySerializer,
    PurchaseOrderSerializer,
    PurchaseOrderListSerializer,
    GoodsReceiptSerializer,
    GoodsReceiptListSerializer,
    VendorBillSerializer,
    VendorBillListSerializer,
    SalesOrderSerializer,
    SalesOrderListSerializer,
    DeliveryNoteSerializer,
    DeliveryNoteListSerializer,
    SalesInvoiceSerializer,
    SalesInvoiceListSerializer,
    PaymentSerializer,
    StockMoveSerializer,
)
from .permissions import (
    IsOwner,
    IsOwnerOrReadOnly,
    IsSameStore,
    IsSalespersonOrOwner,
    CanCreateInvoice,
    CanViewReports,
)

logger = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════════════════════════════
# STORE & ROLE
# ══════════════════════════════════════════════════════════════════════════════


class StoreListView(generics.ListAPIView):
    queryset = Store.objects.filter(is_active=True)
    serializer_class = StoreListSerializer
    permission_classes = []
    pagination_class = None

    @swagger_auto_schema(tags=["Stores"], operation_summary="List active stores")
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)


class StoreDetailView(generics.RetrieveAPIView):
    serializer_class = StoreSerializer
    permission_classes = [permissions.IsAuthenticated]

    @swagger_auto_schema(tags=["Stores"], operation_summary="Get current user's store")
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    def get_object(self):
        return self.request.user.store


class RoleListView(generics.ListAPIView):
    queryset = Role.objects.all().order_by("name")
    serializer_class = RoleSerializer
    permission_classes = []
    pagination_class = None

    @swagger_auto_schema(tags=["Roles"], operation_summary="List all roles")
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)


class RoleDetailView(generics.RetrieveUpdateAPIView):
    """Lets an admin edit a role's permission matrix (Settings → Permissions).
    Roles themselves are a fixed preset (see Role.ROLE_CHOICES) — name and
    display_name aren't editable here, only description/permissions."""
    queryset = Role.objects.all()
    serializer_class = RoleSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_serializer(self, *args, **kwargs):
        serializer = super().get_serializer(*args, **kwargs)
        if self.request.method in ("PUT", "PATCH"):
            for field in ("name", "display_name"):
                serializer.fields[field].read_only = True
        return serializer


# ══════════════════════════════════════════════════════════════════════════════
# CATEGORIES
# ══════════════════════════════════════════════════════════════════════════════


class CategoryListCreateView(generics.ListCreateAPIView):
    serializer_class = CategorySerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["name", "description"]
    ordering_fields = ["name", "created_at"]
    ordering = ["name"]

    @swagger_auto_schema(tags=["Categories"], operation_summary="List categories")
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    @swagger_auto_schema(tags=["Categories"], operation_summary="Create a category")
    def post(self, request, *args, **kwargs):
        return super().post(request, *args, **kwargs)

    def get_queryset(self):
        return Category.objects.filter(
            store=self.request.user.store, is_active=True
        ).select_related("parent", "store")

    def perform_create(self, serializer):
        serializer.save(store=self.request.user.store)


class CategoryDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = CategorySerializer
    permission_classes = [permissions.IsAuthenticated, IsSameStore]

    @swagger_auto_schema(tags=["Categories"], operation_summary="Get category")
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    @swagger_auto_schema(tags=["Categories"], operation_summary="Update category")
    def put(self, request, *args, **kwargs):
        return super().put(request, *args, **kwargs)

    @swagger_auto_schema(
        tags=["Categories"], operation_summary="Partial update category"
    )
    def patch(self, request, *args, **kwargs):
        return super().patch(request, *args, **kwargs)

    @swagger_auto_schema(
        tags=["Categories"], operation_summary="Delete category (soft)"
    )
    def delete(self, request, *args, **kwargs):
        return super().delete(request, *args, **kwargs)

    def get_queryset(self):
        return Category.objects.filter(store=self.request.user.store).select_related(
            "parent", "store"
        )

    def perform_destroy(self, instance):
        instance.is_active = False
        instance.save()


# ══════════════════════════════════════════════════════════════════════════════
# PRODUCTS
# ══════════════════════════════════════════════════════════════════════════════


class ProductListCreateView(generics.ListCreateAPIView):
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = ["category", "is_active"]
    search_fields = ["name", "code", "barcode", "description"]
    ordering_fields = ["name", "price", "stock", "created_at"]
    ordering = ["name"]

    @swagger_auto_schema(
        tags=["Products"],
        operation_summary="List products",
        operation_description=(
            "Returns all products for the authenticated user's store.\n\n"
            "**Extra query params:** `low_stock=true` — return only items at or below threshold."
        ),
    )
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    @swagger_auto_schema(tags=["Products"], operation_summary="Create a product")
    def post(self, request, *args, **kwargs):
        return super().post(request, *args, **kwargs)

    def get_queryset(self):
        queryset = Product.objects.filter(store=self.request.user.store).select_related(
            "category", "store", "created_by"
        )
        if self.request.query_params.get("low_stock") == "true":
            queryset = queryset.filter(stock__lte=F("low_stock_threshold"))
        return queryset

    def get_serializer_class(self):
        return (
            ProductListSerializer if self.request.method == "GET" else ProductSerializer
        )

    def perform_create(self, serializer):
        serializer.save(store=self.request.user.store, created_by=self.request.user)


class ProductDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = ProductSerializer
    permission_classes = [permissions.IsAuthenticated, IsSameStore]

    @swagger_auto_schema(tags=["Products"], operation_summary="Get product")
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    @swagger_auto_schema(tags=["Products"], operation_summary="Update product")
    def put(self, request, *args, **kwargs):
        return super().put(request, *args, **kwargs)

    @swagger_auto_schema(tags=["Products"], operation_summary="Partial update product")
    def patch(self, request, *args, **kwargs):
        return super().patch(request, *args, **kwargs)

    @swagger_auto_schema(tags=["Products"], operation_summary="Delete product (soft)")
    def delete(self, request, *args, **kwargs):
        return super().delete(request, *args, **kwargs)

    def get_queryset(self):
        return Product.objects.filter(store=self.request.user.store).select_related(
            "category", "store", "created_by"
        )

    def perform_destroy(self, instance):
        instance.is_active = False
        instance.save()


class LowStockProductsView(generics.ListAPIView):
    serializer_class = ProductListSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = None

    @swagger_auto_schema(tags=["Products"], operation_summary="List low-stock products")
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    def get_queryset(self):
        return (
            Product.objects.filter(store=self.request.user.store, is_active=True)
            .filter(stock__lte=F("low_stock_threshold"))
            .select_related("category", "store")
            .order_by("stock")
        )


# ══════════════════════════════════════════════════════════════════════════════
# POS INVOICES
# ══════════════════════════════════════════════════════════════════════════════


class InvoiceListCreateView(generics.ListCreateAPIView):
    permission_classes = [permissions.IsAuthenticated, CanCreateInvoice]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ["sync_status"]
    ordering_fields = ["created_at", "total"]
    ordering = ["-created_at"]

    @swagger_auto_schema(
        tags=["Invoices"],
        operation_summary="List invoices",
        operation_description=(
            "Returns invoices for the current store.\n\n"
            "**Salespeople** see only their own invoices.\n"
            "**Owners/Managers** see all store invoices.\n\n"
            "Query params: `sync_status=PENDING|SYNCED|FAILED`, `start_date=YYYY-MM-DD`, `end_date=YYYY-MM-DD`"
        ),
    )
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    @swagger_auto_schema(
        tags=["Invoices"],
        operation_summary="Create invoice (make a sale)",
        operation_description="Creates an invoice and automatically deducts stock for each item.",
    )
    def post(self, request, *args, **kwargs):
        return super().post(request, *args, **kwargs)

    def get_serializer_class(self):
        return (
            InvoiceListSerializer if self.request.method == "GET" else InvoiceSerializer
        )

    def get_queryset(self):
        user = self.request.user
        queryset = Invoice.objects.filter(store=user.store)
        if user.role.name == "salesperson":
            queryset = queryset.filter(salesperson=user)
        start_date = self.request.query_params.get("start_date")
        end_date = self.request.query_params.get("end_date")
        if start_date:
            try:
                queryset = queryset.filter(
                    created_at__gte=datetime.fromisoformat(
                        start_date.replace("Z", "+00:00")
                    )
                )
            except ValueError:
                pass
        if end_date:
            try:
                queryset = queryset.filter(
                    created_at__lte=datetime.fromisoformat(
                        end_date.replace("Z", "+00:00")
                    )
                )
            except ValueError:
                pass
        return queryset.select_related(
            "store", "salesperson", "salesperson__role"
        ).prefetch_related("items__product")

    @transaction.atomic
    def perform_create(self, serializer):
        try:
            invoice = serializer.save(
                store=self.request.user.store, salesperson=self.request.user
            )
            logger.info(
                f"Invoice {invoice.invoice_number} created by {self.request.user.email}"
            )
        except Exception as e:
            logger.error(f"Error creating invoice: {e}")
            raise ValidationError(
                {"error": "Failed to create invoice. Please try again."}
            )


class InvoiceDetailView(generics.RetrieveAPIView):
    serializer_class = InvoiceSerializer
    permission_classes = [permissions.IsAuthenticated, IsSalespersonOrOwner]

    @swagger_auto_schema(
        tags=["Invoices"],
        operation_summary="Get invoice details",
        operation_description="Returns a single invoice with all items. Salespeople can only retrieve their own invoices.",
    )
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    def get_queryset(self):
        user = self.request.user
        queryset = Invoice.objects.filter(store=user.store)
        if user.role.name == "salesperson":
            queryset = queryset.filter(salesperson=user)
        return queryset.select_related(
            "store", "salesperson", "salesperson__role"
        ).prefetch_related("items__product")


class BulkInvoiceSyncView(generics.CreateAPIView):
    serializer_class = BulkInvoiceSyncSerializer
    permission_classes = [permissions.IsAuthenticated, CanCreateInvoice]

    @swagger_auto_schema(
        tags=["Invoices"],
        operation_summary="Bulk sync offline invoices",
        operation_description=(
            "Accepts an array of invoices created while offline and saves them to the server. "
            "Each invoice is processed independently — a failure in one does not roll back others."
        ),
    )
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            result = serializer.save()
            log_status = "completed" if result["failed"] == 0 else "failed"
            SyncLog.objects.create(
                user=request.user,
                store=request.user.store,
                sync_type="invoice",
                status=log_status,
                items_synced=result["synced"],
                items_failed=result["failed"],
                details=result,
                completed_at=timezone.now(),
            )
            return Response(result, status=status.HTTP_200_OK)
        except Exception as e:
            logger.error(f"Bulk sync error: {e}")
            SyncLog.objects.create(
                user=request.user,
                store=request.user.store,
                sync_type="invoice",
                status="failed",
                items_synced=0,
                items_failed=0,
                error_message=str(e),
                completed_at=timezone.now(),
            )
            return Response(
                {"error": "Sync failed. Please try again."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


# ══════════════════════════════════════════════════════════════════════════════
# DASHBOARD & ANALYTICS
# ══════════════════════════════════════════════════════════════════════════════


class DashboardStatsView(views.APIView):
    permission_classes = [permissions.IsAuthenticated, CanViewReports]

    @swagger_auto_schema(
        tags=["Dashboard"],
        operation_summary="Get dashboard statistics",
        operation_description="Returns today's sales, weekly/monthly totals, top product, active salespeople, and low-stock count.",
        responses={
            200: openapi.Response(
                description="Dashboard stats",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        "today_sales": openapi.Schema(type=openapi.TYPE_NUMBER),
                        "invoice_count": openapi.Schema(type=openapi.TYPE_INTEGER),
                        "top_product": openapi.Schema(type=openapi.TYPE_STRING),
                        "active_salespeople": openapi.Schema(type=openapi.TYPE_INTEGER),
                        "week_sales": openapi.Schema(type=openapi.TYPE_NUMBER),
                        "month_sales": openapi.Schema(type=openapi.TYPE_NUMBER),
                        "low_stock_products": openapi.Schema(type=openapi.TYPE_INTEGER),
                    },
                ),
            ),
        },
    )
    def get(self, request):
        try:
            store = request.user.store
            today = timezone.now().date()
            week_ago = today - timedelta(days=7)
            month_ago = today - timedelta(days=30)

            today_invoices = Invoice.objects.filter(
                store=store, created_at__date=today, sync_status="SYNCED"
            )
            today_sales = today_invoices.aggregate(total=Sum("total"))[
                "total"
            ] or Decimal("0.00")
            invoice_count = today_invoices.count()
            active_salespeople = today_invoices.values("salesperson").distinct().count()
            week_sales = Invoice.objects.filter(
                store=store, created_at__date__gte=week_ago, sync_status="SYNCED"
            ).aggregate(total=Sum("total"))["total"] or Decimal("0.00")
            month_sales = Invoice.objects.filter(
                store=store, created_at__date__gte=month_ago, sync_status="SYNCED"
            ).aggregate(total=Sum("total"))["total"] or Decimal("0.00")
            top_product = (
                InvoiceItem.objects.filter(
                    invoice__store=store,
                    invoice__created_at__date=today,
                    invoice__sync_status="SYNCED",
                )
                .values("product_name")
                .annotate(total_quantity=Sum("quantity"))
                .order_by("-total_quantity")
                .first()
            )
            low_stock_count = (
                Product.objects.filter(store=store, is_active=True)
                .filter(stock__lte=F("low_stock_threshold"))
                .count()
            )
            data = {
                "today_sales": today_sales,
                "invoice_count": invoice_count,
                "top_product": top_product["product_name"] if top_product else "N/A",
                "active_salespeople": active_salespeople,
                "week_sales": week_sales,
                "month_sales": month_sales,
                "low_stock_products": low_stock_count,
            }
            return Response(DashboardStatsSerializer(data).data)
        except Exception as e:
            logger.error(f"Dashboard stats error: {e}")
            return Response(
                {"error": "Failed to fetch dashboard statistics"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class SalesReportView(views.APIView):
    permission_classes = [permissions.IsAuthenticated, CanViewReports]

    @swagger_auto_schema(
        tags=["Reports"],
        operation_summary="Sales report by salesperson",
        manual_parameters=[
            openapi.Parameter(
                "start_date", openapi.IN_QUERY, type=openapi.TYPE_STRING, format="date"
            ),
            openapi.Parameter(
                "end_date", openapi.IN_QUERY, type=openapi.TYPE_STRING, format="date"
            ),
        ],
    )
    def get(self, request):
        try:
            store = request.user.store
            start_date = request.query_params.get(
                "start_date", (timezone.now().date() - timedelta(days=30)).isoformat()
            )
            end_date = request.query_params.get(
                "end_date", timezone.now().date().isoformat()
            )
            sales_data = (
                Invoice.objects.filter(
                    store=store,
                    created_at__date__gte=start_date,
                    created_at__date__lte=end_date,
                    sync_status="SYNCED",
                )
                .values("salesperson__id", "salesperson__name")
                .annotate(
                    total_sales=Sum("total"),
                    invoice_count=Count("id"),
                    average_sale=Avg("total"),
                )
                .order_by("-total_sales")
            )
            report_data = [
                {
                    "salesperson_id": item["salesperson__id"],
                    "salesperson_name": item["salesperson__name"],
                    "total_sales": item["total_sales"],
                    "invoice_count": item["invoice_count"],
                    "average_sale": item["average_sale"],
                }
                for item in sales_data
            ]
            return Response(SalesReportSerializer(report_data, many=True).data)
        except Exception as e:
            logger.error(f"Sales report error: {e}")
            return Response(
                {"error": "Failed to generate sales report"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class ProductReportView(views.APIView):
    permission_classes = [permissions.IsAuthenticated, CanViewReports]

    @swagger_auto_schema(
        tags=["Reports"],
        operation_summary="Product sales report",
        manual_parameters=[
            openapi.Parameter(
                "start_date", openapi.IN_QUERY, type=openapi.TYPE_STRING, format="date"
            ),
            openapi.Parameter(
                "end_date", openapi.IN_QUERY, type=openapi.TYPE_STRING, format="date"
            ),
            openapi.Parameter("limit", openapi.IN_QUERY, type=openapi.TYPE_INTEGER),
        ],
    )
    def get(self, request):
        try:
            store = request.user.store
            start_date = request.query_params.get(
                "start_date", (timezone.now().date() - timedelta(days=30)).isoformat()
            )
            end_date = request.query_params.get(
                "end_date", timezone.now().date().isoformat()
            )
            limit = int(request.query_params.get("limit", 20))
            product_data = (
                InvoiceItem.objects.filter(
                    invoice__store=store,
                    invoice__created_at__date__gte=start_date,
                    invoice__created_at__date__lte=end_date,
                    invoice__sync_status="SYNCED",
                )
                .values("product__id", "product__name", "product__code")
                .annotate(quantity_sold=Sum("quantity"), total_revenue=Sum("total"))
                .order_by("-quantity_sold")[:limit]
            )
            report_data = [
                {
                    "product_id": item["product__id"],
                    "product_name": item["product__name"],
                    "product_code": item["product__code"],
                    "quantity_sold": item["quantity_sold"],
                    "total_revenue": item["total_revenue"],
                }
                for item in product_data
            ]
            return Response(ProductReportSerializer(report_data, many=True).data)
        except Exception as e:
            logger.error(f"Product report error: {e}")
            return Response(
                {"error": "Failed to generate product report"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


# ══════════════════════════════════════════════════════════════════════════════
# SYNC
# ══════════════════════════════════════════════════════════════════════════════


class SyncStatusView(views.APIView):
    permission_classes = [permissions.IsAuthenticated]

    @swagger_auto_schema(
        tags=["Sync"],
        operation_summary="Get sync status",
        responses={
            200: openapi.Response(
                description="Sync status",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        "pending_invoices": openapi.Schema(type=openapi.TYPE_INTEGER),
                        "last_sync_time": openapi.Schema(
                            type=openapi.TYPE_STRING, format="date-time"
                        ),
                        "sync_status": openapi.Schema(
                            type=openapi.TYPE_STRING, enum=["online", "pending"]
                        ),
                    },
                ),
            ),
        },
    )
    def get(self, request):
        try:
            user = request.user
            pending_invoices = Invoice.objects.filter(
                salesperson=user, sync_status="PENDING"
            ).count()
            last_sync = (
                SyncLog.objects.filter(user=user, status="completed")
                .order_by("-completed_at")
                .first()
            )
            return Response(
                {
                    "pending_invoices": pending_invoices,
                    "last_sync_time": last_sync.completed_at if last_sync else None,
                    "sync_status": "online" if pending_invoices == 0 else "pending",
                }
            )
        except Exception as e:
            logger.error(f"Sync status error: {e}")
            return Response(
                {"error": "Failed to fetch sync status"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class SyncHistoryView(generics.ListAPIView):
    serializer_class = SyncLogSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = None

    @swagger_auto_schema(tags=["Sync"], operation_summary="Get sync history")
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    def get_queryset(self):
        return (
            SyncLog.objects.filter(user=self.request.user)
            .select_related("store")
            .order_by("-started_at")[:20]
        )


# ══════════════════════════════════════════════════════════════════════════════
# PROFILE
# ══════════════════════════════════════════════════════════════════════════════


class UserProfileView(generics.RetrieveUpdateAPIView):
    serializer_class = UserProfileSerializer
    permission_classes = [permissions.IsAuthenticated]

    @swagger_auto_schema(tags=["Profile"], operation_summary="Get user profile")
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    @swagger_auto_schema(tags=["Profile"], operation_summary="Update user profile")
    def patch(self, request, *args, **kwargs):
        return super().patch(request, *args, **kwargs)

    @swagger_auto_schema(auto_schema=None)
    def put(self, request, *args, **kwargs):
        return super().put(request, *args, **kwargs)

    def get_object(self):
        return self.request.user


# ══════════════════════════════════════════════════════════════════════════════
# CONTACTS
# ══════════════════════════════════════════════════════════════════════════════


class ContactListCreateView(generics.ListCreateAPIView):
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = ["contact_type", "is_active"]
    search_fields = ["name", "email", "phone"]
    ordering = ["name"]

    def get_serializer_class(self):
        return (
            ContactSerializer
            if self.request.method == "POST"
            else ContactListSerializer
        )

    def get_queryset(self):
        return Contact.objects.filter(store=self.request.user.store)

    def perform_create(self, serializer):
        serializer.save(store=self.request.user.store)


class ContactDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = ContactSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return Contact.objects.filter(store=self.request.user.store)

    def perform_destroy(self, instance):
        instance.is_active = False
        instance.save()


# ══════════════════════════════════════════════════════════════════════════════
# ACCOUNTS & JOURNALS
# ══════════════════════════════════════════════════════════════════════════════


class AccountListCreateView(generics.ListCreateAPIView):
    serializer_class = AccountSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [filters.SearchFilter, DjangoFilterBackend]
    filterset_fields = ["account_type", "is_active"]
    search_fields = ["code", "name"]

    def get_queryset(self):
        return Account.objects.filter(store=self.request.user.store)

    def perform_create(self, serializer):
        serializer.save(store=self.request.user.store)


class AccountDetailView(generics.RetrieveUpdateAPIView):
    serializer_class = AccountSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return Account.objects.filter(store=self.request.user.store, is_system=False)


class JournalListView(generics.ListAPIView):
    serializer_class = JournalSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return Journal.objects.filter(store=self.request.user.store)


class JournalEntryListView(generics.ListAPIView):
    serializer_class = JournalEntrySerializer
    permission_classes = [permissions.IsAuthenticated, CanViewReports]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ["status", "source_type"]
    ordering = ["-date"]

    def get_queryset(self):
        return JournalEntry.objects.filter(
            store=self.request.user.store
        ).prefetch_related("lines__account")


class JournalEntryDetailView(generics.RetrieveAPIView):
    serializer_class = JournalEntrySerializer
    permission_classes = [permissions.IsAuthenticated, CanViewReports]

    def get_queryset(self):
        return JournalEntry.objects.filter(store=self.request.user.store)


# ══════════════════════════════════════════════════════════════════════════════
# PURCHASE ORDERS
# ══════════════════════════════════════════════════════════════════════════════


class PurchaseOrderListCreateView(generics.ListCreateAPIView):
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = ["status", "vendor"]
    search_fields = ["po_number", "vendor__name"]
    ordering = ["-created_at"]

    def get_serializer_class(self):
        return (
            PurchaseOrderSerializer
            if self.request.method == "POST"
            else PurchaseOrderListSerializer
        )

    def get_queryset(self):
        return PurchaseOrder.objects.filter(
            store=self.request.user.store
        ).select_related("vendor")


class PurchaseOrderDetailView(generics.RetrieveUpdateAPIView):
    serializer_class = PurchaseOrderSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return PurchaseOrder.objects.filter(store=self.request.user.store)

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        if instance.status not in ("rfq",):
            return Response(
                {"error": "Only RFQ orders can be edited."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        return super().update(request, *args, **kwargs)


class ConfirmPurchaseOrderView(views.APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk):
        try:
            po = PurchaseOrder.objects.get(pk=pk, store=request.user.store)
            po.confirm()
            return Response({"status": "confirmed", "po_number": po.po_number})
        except PurchaseOrder.DoesNotExist:
            return Response({"error": "Not found"}, status=status.HTTP_404_NOT_FOUND)
        except ValueError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


# ══════════════════════════════════════════════════════════════════════════════
# GOODS RECEIPTS
# ══════════════════════════════════════════════════════════════════════════════


class GoodsReceiptListCreateView(generics.ListCreateAPIView):
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ["status", "purchase_order"]
    ordering = ["-received_date"]

    def get_serializer_class(self):
        return (
            GoodsReceiptSerializer
            if self.request.method == "POST"
            else GoodsReceiptListSerializer
        )

    def get_queryset(self):
        return GoodsReceipt.objects.filter(
            store=self.request.user.store
        ).select_related("purchase_order__vendor")


class GoodsReceiptDetailView(generics.RetrieveAPIView):
    serializer_class = GoodsReceiptSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return GoodsReceipt.objects.filter(store=self.request.user.store)


class ValidateGoodsReceiptView(views.APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk):
        try:
            receipt = GoodsReceipt.objects.get(pk=pk, store=request.user.store)
            if receipt.status != "draft":
                return Response(
                    {"error": "Only draft receipts can be validated."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            _post_goods_receipt(receipt, request.user)
            return Response(
                {"status": "validated", "receipt_number": receipt.receipt_number}
            )
        except GoodsReceipt.DoesNotExist:
            return Response({"error": "Not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Goods receipt validation error: {e}", exc_info=True)
            return Response(
                {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# ══════════════════════════════════════════════════════════════════════════════
# VENDOR BILLS
# ══════════════════════════════════════════════════════════════════════════════


class VendorBillListCreateView(generics.ListCreateAPIView):
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = ["status", "vendor"]
    search_fields = ["bill_number", "vendor__name"]
    ordering = ["-bill_date"]

    def get_serializer_class(self):
        return (
            VendorBillSerializer
            if self.request.method == "POST"
            else VendorBillListSerializer
        )

    def get_queryset(self):
        return VendorBill.objects.filter(store=self.request.user.store).select_related(
            "vendor"
        )


class VendorBillDetailView(generics.RetrieveAPIView):
    serializer_class = VendorBillSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return VendorBill.objects.filter(store=self.request.user.store)


class ConfirmVendorBillView(views.APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk):
        try:
            bill = VendorBill.objects.get(pk=pk, store=request.user.store)
            if bill.status != "draft":
                return Response(
                    {"error": "Only draft bills can be confirmed."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            _post_vendor_bill(bill, request.user)
            return Response({"status": "confirmed", "bill_number": bill.bill_number})
        except VendorBill.DoesNotExist:
            return Response({"error": "Not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Vendor bill confirmation error: {e}", exc_info=True)
            return Response(
                {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# ══════════════════════════════════════════════════════════════════════════════
# SALES ORDERS
# ══════════════════════════════════════════════════════════════════════════════


class SalesOrderListCreateView(generics.ListCreateAPIView):
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = ["status", "customer"]
    search_fields = ["so_number", "customer__name"]
    ordering = ["-created_at"]

    def get_serializer_class(self):
        return (
            SalesOrderSerializer
            if self.request.method == "POST"
            else SalesOrderListSerializer
        )

    def get_queryset(self):
        return SalesOrder.objects.filter(store=self.request.user.store).select_related(
            "customer", "salesperson"
        )


class SalesOrderDetailView(generics.RetrieveUpdateAPIView):
    serializer_class = SalesOrderSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return SalesOrder.objects.filter(store=self.request.user.store)

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        if instance.status not in ("draft",):
            return Response(
                {"error": "Only draft sales orders can be edited."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        return super().update(request, *args, **kwargs)


class ConfirmSalesOrderView(views.APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk):
        try:
            so = SalesOrder.objects.get(pk=pk, store=request.user.store)
            so.confirm()
            return Response({"status": "confirmed", "so_number": so.so_number})
        except SalesOrder.DoesNotExist:
            return Response({"error": "Not found"}, status=status.HTTP_404_NOT_FOUND)
        except ValueError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


# ══════════════════════════════════════════════════════════════════════════════
# DELIVERY NOTES
# ══════════════════════════════════════════════════════════════════════════════


class DeliveryNoteListCreateView(generics.ListCreateAPIView):
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ["status", "sales_order"]
    ordering = ["-delivery_date"]

    def get_serializer_class(self):
        return (
            DeliveryNoteSerializer
            if self.request.method == "POST"
            else DeliveryNoteListSerializer
        )

    def get_queryset(self):
        return DeliveryNote.objects.filter(
            store=self.request.user.store
        ).select_related("sales_order__customer")


class DeliveryNoteDetailView(generics.RetrieveAPIView):
    serializer_class = DeliveryNoteSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return DeliveryNote.objects.filter(store=self.request.user.store)


class ValidateDeliveryNoteView(views.APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk):
        try:
            note = DeliveryNote.objects.get(pk=pk, store=request.user.store)
            if note.status != "draft":
                return Response(
                    {"error": "Only draft delivery notes can be validated."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            _validate_delivery_note(note, request.user)
            return Response(
                {"status": "delivered", "delivery_number": note.delivery_number}
            )
        except DeliveryNote.DoesNotExist:
            return Response({"error": "Not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Delivery note validation error: {e}", exc_info=True)
            return Response(
                {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# ══════════════════════════════════════════════════════════════════════════════
# SALES INVOICES
# ══════════════════════════════════════════════════════════════════════════════


class SalesInvoiceListCreateView(generics.ListCreateAPIView):
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = ["status", "customer"]
    search_fields = ["invoice_number", "customer__name"]
    ordering = ["-invoice_date"]

    def get_serializer_class(self):
        return (
            SalesInvoiceSerializer
            if self.request.method == "POST"
            else SalesInvoiceListSerializer
        )

    def get_queryset(self):
        return SalesInvoice.objects.filter(
            store=self.request.user.store
        ).select_related("customer", "salesperson")


class SalesInvoiceDetailView(generics.RetrieveAPIView):
    serializer_class = SalesInvoiceSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return SalesInvoice.objects.filter(store=self.request.user.store)


class PostSalesInvoiceView(views.APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk):
        try:
            invoice = SalesInvoice.objects.get(pk=pk, store=request.user.store)
            if invoice.status not in ("draft",):
                return Response(
                    {"error": "Only draft invoices can be posted."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            _post_sales_invoice(invoice, request.user)
            return Response(
                {"status": invoice.status, "invoice_number": invoice.invoice_number}
            )
        except SalesInvoice.DoesNotExist:
            return Response({"error": "Not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Sales invoice post error: {e}", exc_info=True)
            return Response(
                {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# ══════════════════════════════════════════════════════════════════════════════
# PAYMENTS
# ══════════════════════════════════════════════════════════════════════════════


class PaymentListCreateView(generics.ListCreateAPIView):
    serializer_class = PaymentSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ["payment_type", "status", "payment_method"]
    ordering = ["-date"]

    def get_queryset(self):
        return Payment.objects.filter(store=self.request.user.store).select_related(
            "contact"
        )


class PaymentDetailView(generics.RetrieveAPIView):
    serializer_class = PaymentSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return Payment.objects.filter(store=self.request.user.store)


class ConfirmPaymentView(views.APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk):
        try:
            payment = Payment.objects.get(pk=pk, store=request.user.store)
            if payment.status != "draft":
                return Response(
                    {"error": "Only draft payments can be confirmed."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            _post_payment(payment, request.user)
            return Response(
                {"status": "confirmed", "payment_number": payment.payment_number}
            )
        except Payment.DoesNotExist:
            return Response({"error": "Not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Payment confirmation error: {e}", exc_info=True)
            return Response(
                {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# ══════════════════════════════════════════════════════════════════════════════
# STOCK & REPORTING
# ══════════════════════════════════════════════════════════════════════════════


class StockMoveListView(generics.ListAPIView):
    serializer_class = StockMoveSerializer
    permission_classes = [permissions.IsAuthenticated, CanViewReports]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ["move_type", "product"]
    ordering = ["-date"]

    def get_queryset(self):
        return StockMove.objects.filter(store=self.request.user.store).select_related(
            "product"
        )


class TrialBalanceView(views.APIView):
    permission_classes = [permissions.IsAuthenticated, CanViewReports]

    def get(self, request):
        from django.db.models import Sum as DSum

        store = request.user.store
        lines = (
            JournalEntryLine.objects.filter(
                journal_entry__store=store, journal_entry__status="posted"
            )
            .values("account__code", "account__name", "account__account_type")
            .annotate(total_debit=DSum("debit"), total_credit=DSum("credit"))
            .order_by("account__code")
        )
        rows = [
            {
                "account_code": l["account__code"],
                "account_name": l["account__name"],
                "account_type": l["account__account_type"],
                "total_debit": l["total_debit"] or Decimal("0"),
                "total_credit": l["total_credit"] or Decimal("0"),
                "net_balance": (l["total_debit"] or Decimal("0"))
                - (l["total_credit"] or Decimal("0")),
            }
            for l in lines
        ]
        return Response({"trial_balance": rows})


# ══════════════════════════════════════════════════════════════════════════════
# ADMIN / EXCEL VIEWS
# ══════════════════════════════════════════════════════════════════════════════


@staff_member_required
def download_product_template(request):
    if not request.user.is_superuser and not request.user.store:
        messages.error(
            request, "You must be assigned to a store to download the template."
        )
        return redirect("admin:pos_app_product_changelist")

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    ws_instructions = wb.create_sheet("Instructions", 0)
    ws_instructions["A1"] = "Bulk Product Upload Instructions"
    ws_instructions["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws_instructions["A1"].fill = PatternFill(
        start_color="2563EB", end_color="2563EB", fill_type="solid"
    )
    ws_instructions.merge_cells("A1:D1")
    ws_instructions.row_dimensions[1].height = 30
    ws_instructions["A1"].alignment = Alignment(horizontal="center", vertical="center")

    instructions = [
        ("", ""),
        ("STEP-BY-STEP GUIDE", ""),
        (
            "1. Fill Product Data",
            'Go to the "Products" sheet and fill in your product information',
        ),
        ("2. Follow Format", "Each column has specific requirements (see below)"),
        ("3. Check Categories", 'Use category names from the "Categories" sheet'),
        ("4. Save File", "Save the file as .xlsx format"),
        ("5. Upload", "Go back to admin and upload the file"),
        ("", ""),
        ("COLUMN DESCRIPTIONS", ""),
        ("code *", "Unique product code/SKU (REQUIRED)"),
        ("name *", "Product name (REQUIRED)"),
        ("description", "Product description (OPTIONAL)"),
        ("category_name", "Category name — must match existing category"),
        ("price *", "Selling price (REQUIRED)"),
        ("cost", "Cost price (OPTIONAL)"),
        ("stock *", "Initial stock quantity (REQUIRED)"),
        ("low_stock_threshold", "Alert threshold (OPTIONAL, default: 10)"),
        ("barcode", "Product barcode (OPTIONAL)"),
        ("image_url", "Product image URL (OPTIONAL)"),
        ("", ""),
        ("IMPORTANT NOTES", ""),
        ("✓", "Fields marked with * are REQUIRED"),
        ("✓", "Code and barcode must be UNIQUE"),
        ("✓", "Category name must match exactly"),
    ]
    for idx, (col1, col2) in enumerate(instructions, start=2):
        ws_instructions[f"A{idx}"] = col1
        ws_instructions[f"B{idx}"] = col2
        if col1 in ("STEP-BY-STEP GUIDE", "COLUMN DESCRIPTIONS", "IMPORTANT NOTES"):
            ws_instructions[f"A{idx}"].font = Font(bold=True, size=12, color="1F2937")
            ws_instructions[f"A{idx}"].fill = PatternFill(
                start_color="F3F4F6", end_color="F3F4F6", fill_type="solid"
            )
            ws_instructions.merge_cells(f"A{idx}:B{idx}")
    ws_instructions.column_dimensions["A"].width = 25
    ws_instructions.column_dimensions["B"].width = 70

    ws_products = wb.create_sheet("Products", 1)
    headers = [
        "code",
        "name",
        "description",
        "category_name",
        "price",
        "cost",
        "stock",
        "low_stock_threshold",
        "barcode",
        "image_url",
    ]
    header_fill = PatternFill(
        start_color="16A34A", end_color="16A34A", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for col_num, header in enumerate(headers, 1):
        cell = ws_products.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    sample_products = [
        [
            "LAPTOP001",
            "Dell XPS 15",
            "High-performance laptop",
            "Electronics",
            "1500000",
            "1200000",
            "25",
            "5",
            "123456789001",
            "",
        ],
        [
            "PHONE001",
            "Samsung Galaxy S23",
            "Latest smartphone",
            "Electronics",
            "800000",
            "650000",
            "50",
            "10",
            "123456789002",
            "",
        ],
    ]
    for row_num, product in enumerate(sample_products, 2):
        for col_num, value in enumerate(product, 1):
            cell = ws_products.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin_border
            cell.fill = PatternFill(
                start_color="F9FAFB", end_color="F9FAFB", fill_type="solid"
            )

    for col, width in zip("ABCDEFGHIJ", [15, 30, 40, 20, 12, 12, 10, 20, 18, 35]):
        ws_products.column_dimensions[col].width = width
    ws_products.freeze_panes = "A2"

    ws_categories = wb.create_sheet("Categories", 2)
    ws_categories["A1"] = "Available Categories"
    ws_categories["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws_categories["A1"].fill = PatternFill(
        start_color="0891B2", end_color="0891B2", fill_type="solid"
    )
    ws_categories["B1"] = "Store"
    ws_categories["B1"].font = ws_categories["A1"].font
    ws_categories["B1"].fill = ws_categories["A1"].fill
    categories = (
        Category.objects.filter(is_active=True)
        .select_related("store")
        .order_by("store__name", "name")
        if request.user.is_superuser
        else Category.objects.filter(is_active=True, store=request.user.store)
        .select_related("store")
        .order_by("name")
    )
    for idx, cat in enumerate(categories, 2):
        ws_categories[f"A{idx}"] = cat.name
        ws_categories[f"B{idx}"] = cat.store.name if cat.store else "N/A"
    ws_categories.column_dimensions["A"].width = 30
    ws_categories.column_dimensions["B"].width = 25

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        "attachment; filename=product_upload_template.xlsx"
    )
    wb.save(response)
    return response


@staff_member_required
def bulk_upload_products(request):
    if not request.user.is_superuser and not request.user.store:
        messages.error(request, "You must be assigned to a store to upload products.")
        return redirect("admin:pos_app_product_changelist")

    if request.method == "POST":
        excel_file = request.FILES.get("excel_file")
        if not excel_file:
            messages.error(request, "Please select an Excel file to upload.")
            return redirect("admin:pos_app_product_bulk_upload")
        if not excel_file.name.endswith((".xlsx", ".xls")):
            messages.error(
                request, "Invalid file format. Please upload .xlsx or .xls file."
            )
            return redirect("admin:pos_app_product_bulk_upload")
        try:
            wb = openpyxl.load_workbook(excel_file)
            if "Products" not in wb.sheetnames:
                messages.error(request, 'Invalid template. "Products" sheet not found.')
                return redirect("admin:pos_app_product_bulk_upload")
            ws = wb["Products"]
            headers = [cell.value for cell in ws[1]]
            for h in ["code", "name", "price", "stock"]:
                if h not in headers:
                    messages.error(request, f"Missing required column: {h}")
                    return redirect("admin:pos_app_product_bulk_upload")
            user_store = request.user.store
            products_to_create = []
            errors = []
            success_count = 0
            skip_count = 0
            with transaction.atomic():
                for idx, row in enumerate(
                    ws.iter_rows(min_row=2, values_only=True), start=2
                ):
                    if not any(row):
                        skip_count += 1
                        continue
                    row_data = dict(zip(headers, row))
                    if not row_data.get("code"):
                        errors.append(f"Row {idx}: Missing product code")
                        continue
                    if not row_data.get("name"):
                        errors.append(f"Row {idx}: Missing product name")
                        continue
                    if not row_data.get("price"):
                        errors.append(f"Row {idx}: Missing price")
                        continue
                    if Product.objects.filter(
                        code=row_data["code"], store=user_store
                    ).exists():
                        errors.append(
                            f"Row {idx}: Product code '{row_data['code']}' already exists"
                        )
                        continue
                    if (
                        row_data.get("barcode")
                        and Product.objects.filter(
                            barcode=row_data["barcode"], store=user_store
                        ).exists()
                    ):
                        errors.append(
                            f"Row {idx}: Barcode '{row_data['barcode']}' already exists"
                        )
                        continue
                    category = None
                    if row_data.get("category_name"):
                        try:
                            category = Category.objects.get(
                                name=row_data["category_name"],
                                store=user_store,
                                is_active=True,
                            )
                        except Category.DoesNotExist:
                            errors.append(
                                f"Row {idx}: Category '{row_data['category_name']}' not found."
                            )
                            continue
                    try:
                        product = Product(
                            store=user_store,
                            code=row_data["code"],
                            name=row_data["name"],
                            description=row_data.get("description", ""),
                            category=category,
                            price=Decimal(str(row_data["price"])),
                            cost=(
                                Decimal(str(row_data["cost"]))
                                if row_data.get("cost")
                                else None
                            ),
                            stock=int(row_data.get("stock", 0)),
                            low_stock_threshold=int(
                                row_data.get("low_stock_threshold", 10)
                            ),
                            barcode=row_data.get("barcode", ""),
                            image_url=row_data.get("image_url", ""),
                            created_by=request.user,
                            is_active=True,
                        )
                        product.full_clean()
                        products_to_create.append(product)
                        success_count += 1
                    except Exception as e:
                        errors.append(f"Row {idx}: {e}")
                if products_to_create:
                    Product.objects.bulk_create(products_to_create)
            if success_count > 0:
                messages.success(
                    request, f"Successfully uploaded {success_count} product(s)!"
                )
            if skip_count > 0:
                messages.info(request, f"Skipped {skip_count} empty row(s)")
            if errors:
                err_msg = f"Failed to upload {len(errors)} product(s):\n" + "\n".join(
                    errors[:10]
                )
                if len(errors) > 10:
                    err_msg += f"\n... and {len(errors) - 10} more errors"
                messages.warning(request, err_msg)
            return redirect("admin:pos_app_product_changelist")
        except Exception as e:
            logger.error(f"Bulk upload error: {e}", exc_info=True)
            messages.error(request, f"Error processing file: {e}")
            return redirect("admin:pos_app_product_bulk_upload")

    return render(request, "admin/pos_app/product/bulk_upload.html")


@staff_member_required
def export_products_excel(request):
    if not request.user.is_superuser and not request.user.store:
        messages.error(request, "You must be assigned to a store to export products.")
        return redirect("admin:pos_app_product_changelist")
    products = (
        Product.objects.all()
        .select_related("category", "store")
        .order_by("store__name", "code")
        if request.user.is_superuser
        else Product.objects.filter(store=request.user.store)
        .select_related("category", "store")
        .order_by("code")
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    headers = [
        "Code",
        "Name",
        "Description",
        "Category",
        "Price",
        "Cost",
        "Stock",
        "Low Stock Threshold",
        "Barcode",
        "Image URL",
        "Is Active",
        "Created At",
    ]
    header_fill = PatternFill(
        start_color="2563EB", end_color="2563EB", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_num, p in enumerate(products, 2):
        ws.cell(row=row_num, column=1, value=p.code)
        ws.cell(row=row_num, column=2, value=p.name)
        ws.cell(row=row_num, column=3, value=p.description)
        ws.cell(row=row_num, column=4, value=p.category.name if p.category else "")
        ws.cell(row=row_num, column=5, value=float(p.price))
        ws.cell(row=row_num, column=6, value=float(p.cost) if p.cost else "")
        ws.cell(row=row_num, column=7, value=float(p.stock))
        ws.cell(row=row_num, column=8, value=p.low_stock_threshold)
        ws.cell(row=row_num, column=9, value=p.barcode)
        ws.cell(row=row_num, column=10, value=p.image_url)
        ws.cell(row=row_num, column=11, value="Yes" if p.is_active else "No")
        ws.cell(
            row=row_num, column=12, value=p.created_at.strftime("%Y-%m-%d %H:%M:%S")
        )
    for idx, width in enumerate([15, 30, 40, 20, 12, 12, 10, 20, 18, 35, 12, 20], 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    store_code = (
        request.user.store.code
        if not request.user.is_superuser and request.user.store
        else "all"
    )
    response["Content-Disposition"] = (
        f"attachment; filename=products_export_{store_code}.xlsx"
    )
    wb.save(response)
    return response


# ══════════════════════════════════════════════════════════════════════════════
# UTILITY
# ══════════════════════════════════════════════════════════════════════════════


@swagger_auto_schema(
    method="get",
    tags=["Utility"],
    operation_summary="Tenant health check",
    operation_description="Verifies the API is reachable and the user is authenticated.",
    responses={
        200: openapi.Response(
            description="Healthy",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "status": openapi.Schema(
                        type=openapi.TYPE_STRING, example="healthy"
                    ),
                    "timestamp": openapi.Schema(
                        type=openapi.TYPE_STRING, format="date-time"
                    ),
                    "user": openapi.Schema(
                        type=openapi.TYPE_OBJECT,
                        properties={
                            "email": openapi.Schema(type=openapi.TYPE_STRING),
                            "name": openapi.Schema(type=openapi.TYPE_STRING),
                            "store": openapi.Schema(type=openapi.TYPE_STRING),
                            "role": openapi.Schema(type=openapi.TYPE_STRING),
                        },
                    ),
                },
            ),
        ),
    },
)
@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated])
def health_check(request):
    return Response(
        {
            "status": "healthy",
            "timestamp": timezone.now().isoformat(),
            "user": {
                "email": request.user.email,
                "name": request.user.name,
                "store": request.user.store.name if request.user.store else None,
                "role": request.user.role.name if request.user.role else None,
            },
        }
    )


# ══════════════════════════════════════════════════════════════════════════════
# ACCOUNTING BUSINESS LOGIC HELPERS
# ══════════════════════════════════════════════════════════════════════════════


def _get_journal(store, journal_type):
    return Journal.objects.filter(
        store=store, journal_type=journal_type, is_active=True
    ).first()


def _get_account(store, code):
    return Account.objects.filter(store=store, code=code, is_active=True).first()


@transaction.atomic
def _post_goods_receipt(receipt, user):
    store = receipt.store
    je = JournalEntry.objects.create(
        store=store,
        journal=_get_journal(store, "inventory"),
        reference=receipt.receipt_number,
        date=receipt.received_date,
        memo=f"Goods receipt {receipt.receipt_number} from {receipt.purchase_order.vendor.name}",
        source_type="goods_receipt",
        source_id=receipt.id,
        created_by=user,
    )
    ap_account = _get_account(store, "2000")
    for line in receipt.lines.select_related("product", "product__category").all():
        if line.received_qty <= 0:
            continue
        product = line.product
        qty = line.received_qty
        unit_cost = line.unit_cost
        cost_method = product.category.cost_method if product.category else "avco"
        if cost_method == "standard":
            unit_cost = product.standard_cost or unit_cost
        if cost_method in ("avco", "standard"):
            product.update_avco(qty, unit_cost)
        else:  # fifo / lifo
            product.update_avco(qty, unit_cost)
            StockLayer.objects.create(
                store=store,
                product=product,
                quantity_in=qty,
                quantity_remaining=qty,
                unit_cost=unit_cost,
                source_type="purchase_receipt",
                source_id=receipt.id,
                date=receipt.received_date,
            )
        product.cost = unit_cost
        product.save(update_fields=["stock", "stock_value", "avg_cost", "cost"])
        line.purchase_order_line.received_qty += qty
        line.purchase_order_line.save(update_fields=["received_qty"])
        valuation = qty * unit_cost
        StockMove.objects.create(
            store=store,
            product=product,
            move_type="purchase_receipt",
            quantity=qty,
            unit_cost=unit_cost,
            valuation_amount=valuation,
            source_type="goods_receipt",
            source_id=receipt.id,
            journal_entry=je,
            date=receipt.received_date,
        )
        inv_account = (
            product.category.inventory_account if product.category else None
        ) or _get_account(store, "1200")
        JournalEntryLine.objects.create(
            journal_entry=je,
            account=inv_account,
            description=f"{product.name} x{qty}",
            debit=valuation,
            credit=Decimal("0"),
            partner=receipt.purchase_order.vendor,
        )
        JournalEntryLine.objects.create(
            journal_entry=je,
            account=ap_account,
            description=f"{product.name} x{qty}",
            debit=Decimal("0"),
            credit=valuation,
            partner=receipt.purchase_order.vendor,
        )
    je.post()
    po = receipt.purchase_order
    all_lines = po.lines.all()
    po.status = (
        "received"
        if all(l.received_qty >= l.quantity for l in all_lines)
        else "partial"
    )
    po.save(update_fields=["status"])
    receipt.status = "validated"
    receipt.validated_by = user
    receipt.validated_at = timezone.now()
    receipt.journal_entry = je
    receipt.save(
        update_fields=["status", "validated_by", "validated_at", "journal_entry"]
    )


@transaction.atomic
def _post_vendor_bill(bill, user):
    store = bill.store
    je = JournalEntry.objects.create(
        store=store,
        journal=_get_journal(store, "purchase"),
        reference=bill.bill_number,
        date=bill.bill_date,
        memo=f"Vendor bill {bill.bill_number} — {bill.vendor.name}",
        source_type="vendor_bill",
        source_id=bill.id,
        created_by=user,
    )
    ap_account = _get_account(store, "2000")
    for line in bill.lines.select_related("product", "account").all():
        line_account = (
            line.account
            or (
                line.product.category.inventory_account
                if line.product and line.product.category
                else None
            )
            or _get_account(store, "6000")
        )
        desc = line.description or (line.product.name if line.product else "")
        JournalEntryLine.objects.create(
            journal_entry=je,
            account=line_account,
            description=desc,
            debit=line.total,
            credit=Decimal("0"),
            partner=bill.vendor,
        )
        JournalEntryLine.objects.create(
            journal_entry=je,
            account=ap_account,
            description=desc,
            debit=Decimal("0"),
            credit=line.total,
            partner=bill.vendor,
        )
    je.post()
    bill.status = "confirmed"
    bill.journal_entry = je
    bill.save(update_fields=["status", "journal_entry"])


def _consume_fifo(store, product, qty):
    qty = Decimal(str(qty))
    remaining = qty
    total_cost = Decimal("0")
    for layer in StockLayer.objects.filter(
        store=store, product=product, quantity_remaining__gt=0
    ).order_by("date", "created_at"):
        if remaining <= 0:
            break
        consume = min(layer.quantity_remaining, remaining)
        total_cost += consume * layer.unit_cost
        layer.quantity_remaining -= consume
        layer.save(update_fields=["quantity_remaining"])
        remaining -= consume
    product.consume_stock(qty)
    return total_cost / qty if qty else Decimal("0")


def _consume_lifo(store, product, qty):
    qty = Decimal(str(qty))
    remaining = qty
    total_cost = Decimal("0")
    for layer in StockLayer.objects.filter(
        store=store, product=product, quantity_remaining__gt=0
    ).order_by("-date", "-created_at"):
        if remaining <= 0:
            break
        consume = min(layer.quantity_remaining, remaining)
        total_cost += consume * layer.unit_cost
        layer.quantity_remaining -= consume
        layer.save(update_fields=["quantity_remaining"])
        remaining -= consume
    product.consume_stock(qty)
    return total_cost / qty if qty else Decimal("0")


@transaction.atomic
def _post_sales_invoice(invoice, user):
    store = invoice.store
    default_revenue = _get_account(store, "4000")
    default_cogs = _get_account(store, "5000")
    default_inv_asset = _get_account(store, "1200")
    revenue_je = JournalEntry.objects.create(
        store=store,
        journal=_get_journal(store, "sale"),
        reference=invoice.invoice_number,
        date=invoice.invoice_date,
        memo=f"Sales invoice {invoice.invoice_number} — {invoice.customer.name}",
        source_type="sales_invoice",
        source_id=invoice.id,
        created_by=user,
    )
    cogs_je = JournalEntry.objects.create(
        store=store,
        journal=_get_journal(store, "inventory"),
        reference=f"{invoice.invoice_number}-COGS",
        date=invoice.invoice_date,
        memo=f"COGS for invoice {invoice.invoice_number}",
        source_type="sales_invoice",
        source_id=invoice.id,
        created_by=user,
    )
    ar_account = _get_account(store, "1100")
    JournalEntryLine.objects.create(
        journal_entry=revenue_je,
        account=ar_account,
        description=f"Invoice {invoice.invoice_number}",
        debit=invoice.total,
        credit=Decimal("0"),
        partner=invoice.customer,
    )
    for line in invoice.lines.select_related(
        "product", "product__category", "income_account", "cogs_account"
    ).all():
        revenue_account = line.income_account or default_revenue
        cogs_account = line.cogs_account or default_cogs
        JournalEntryLine.objects.create(
            journal_entry=revenue_je,
            account=revenue_account,
            description=f"{line.product.name if line.product else line.description} x{line.quantity}",
            debit=Decimal("0"),
            credit=line.total,
            partner=invoice.customer,
        )
        if line.product:
            product = line.product
            cost_method = product.category.cost_method if product.category else "avco"
            if cost_method == "fifo":
                unit_cost = _consume_fifo(store, product, line.quantity)
            elif cost_method == "lifo":
                unit_cost = _consume_lifo(store, product, line.quantity)
            else:
                unit_cost = (
                    product.standard_cost
                    if cost_method == "standard"
                    else product.avg_cost
                )
                product.consume_stock(line.quantity)
                product.save(update_fields=["stock", "stock_value"])
            cogs_value = Decimal(str(line.quantity)) * unit_cost
            inv_asset_account = (
                product.category.inventory_account if product.category else None
            ) or default_inv_asset
            JournalEntryLine.objects.create(
                journal_entry=cogs_je,
                account=cogs_account,
                description=f"COGS {product.code} x{line.quantity}",
                debit=cogs_value,
                credit=Decimal("0"),
            )
            JournalEntryLine.objects.create(
                journal_entry=cogs_je,
                account=inv_asset_account,
                description=f"Inventory {product.code} x{line.quantity}",
                debit=Decimal("0"),
                credit=cogs_value,
            )
            StockMove.objects.create(
                store=store,
                product=product,
                move_type="sales_delivery",
                quantity=line.quantity,
                unit_cost=unit_cost,
                valuation_amount=cogs_value,
                source_type="sales_invoice",
                source_id=invoice.id,
                journal_entry=cogs_je,
                date=invoice.invoice_date,
            )
            line.cost_at_sale = unit_cost
            line.save(update_fields=["cost_at_sale"])
    revenue_je.post()
    cogs_je.post()
    invoice.status = "sent"
    invoice.journal_entry = revenue_je
    if not invoice.due_date:
        import datetime as _dt

        invoice.due_date = invoice.invoice_date + _dt.timedelta(
            days=invoice.customer.payment_terms_days
        )
    invoice.save(update_fields=["status", "journal_entry", "due_date"])


@transaction.atomic
def _validate_delivery_note(note, user):
    store = note.store
    default_cogs = _get_account(store, "5000")
    default_inv = _get_account(store, "1200")
    je = JournalEntry.objects.create(
        store=store,
        journal=_get_journal(store, "inventory"),
        reference=note.delivery_number,
        date=note.delivery_date,
        memo=f"Delivery {note.delivery_number} for {note.sales_order.customer.name}",
        source_type="delivery_note",
        source_id=note.id,
        created_by=user,
    )
    for line in note.lines.select_related("product", "product__category").all():
        if line.delivered_qty <= 0:
            continue
        product = line.product
        qty = line.delivered_qty
        cost_method = product.category.cost_method if product.category else "avco"
        unit_cost = product.avg_cost
        if cost_method == "fifo":
            unit_cost = _consume_fifo(store, product, qty)
        elif cost_method == "lifo":
            unit_cost = _consume_lifo(store, product, qty)
        else:
            product.consume_stock(qty)
            product.save(update_fields=["stock", "stock_value"])
        cogs_value = Decimal(str(qty)) * unit_cost
        cogs_account = (
            product.category.cogs_account if product.category else None
        ) or default_cogs
        inv_account = (
            product.category.inventory_account if product.category else None
        ) or default_inv
        JournalEntryLine.objects.create(
            journal_entry=je,
            account=cogs_account,
            description=f"COGS {product.code} x{qty}",
            debit=cogs_value,
            credit=Decimal("0"),
        )
        JournalEntryLine.objects.create(
            journal_entry=je,
            account=inv_account,
            description=f"Inventory {product.code} x{qty}",
            debit=Decimal("0"),
            credit=cogs_value,
        )
        StockMove.objects.create(
            store=store,
            product=product,
            move_type="sales_delivery",
            quantity=qty,
            unit_cost=unit_cost,
            valuation_amount=cogs_value,
            source_type="delivery_note",
            source_id=note.id,
            journal_entry=je,
            date=note.delivery_date,
        )
        line.unit_cost_at_delivery = unit_cost
        line.save(update_fields=["unit_cost_at_delivery"])
        line.sales_order_line.delivered_qty += qty
        line.sales_order_line.save(update_fields=["delivered_qty"])
    je.post()
    so = note.sales_order
    all_so_lines = so.lines.all()
    so.status = (
        "delivered"
        if all(l.delivered_qty >= l.quantity for l in all_so_lines)
        else "partial"
    )
    so.save(update_fields=["status"])
    note.status = "delivered"
    note.delivered_by = user
    note.delivered_at = timezone.now()
    note.journal_entry = je
    note.save(update_fields=["status", "delivered_by", "delivered_at", "journal_entry"])


@transaction.atomic
def _post_payment(payment, user):
    store = payment.store
    method_to_journal = {
        "cash": "cash",
        "bank": "bank",
        "card": "bank",
        "cheque": "bank",
        "other": "cash",
    }
    je = JournalEntry.objects.create(
        store=store,
        journal=_get_journal(
            store, method_to_journal.get(payment.payment_method, "cash")
        ),
        reference=payment.payment_number,
        date=payment.date,
        memo=f"Payment {payment.payment_number} — {payment.contact.name}",
        source_type="payment",
        source_id=payment.id,
        created_by=user,
    )
    ar_account = _get_account(store, "1100")
    ap_account = _get_account(store, "2000")
    cash_account = _get_account(
        store, "1000" if payment.payment_method == "cash" else "1010"
    )
    if payment.payment_type == "inbound":
        JournalEntryLine.objects.create(
            journal_entry=je,
            account=cash_account,
            description="Customer payment received",
            debit=payment.amount,
            credit=Decimal("0"),
            partner=payment.contact,
        )
        JournalEntryLine.objects.create(
            journal_entry=je,
            account=ar_account,
            description="Clear accounts receivable",
            debit=Decimal("0"),
            credit=payment.amount,
            partner=payment.contact,
        )
        if payment.sales_invoice:
            inv = payment.sales_invoice
            inv.amount_paid = min(inv.total, inv.amount_paid + payment.amount)
            inv.status = "paid" if inv.amount_paid >= inv.total else "partial"
            inv.save(update_fields=["amount_paid", "status"])
    else:
        JournalEntryLine.objects.create(
            journal_entry=je,
            account=ap_account,
            description="Clear accounts payable",
            debit=payment.amount,
            credit=Decimal("0"),
            partner=payment.contact,
        )
        JournalEntryLine.objects.create(
            journal_entry=je,
            account=cash_account,
            description="Vendor payment sent",
            debit=Decimal("0"),
            credit=payment.amount,
            partner=payment.contact,
        )
        if payment.vendor_bill:
            bill = payment.vendor_bill
            bill.amount_paid = min(bill.total, bill.amount_paid + payment.amount)
            bill.status = "paid" if bill.amount_paid >= bill.total else "partial"
            bill.save(update_fields=["amount_paid", "status"])
    je.post()
    payment.status = "confirmed"
    payment.journal_entry = je
    payment.save(update_fields=["status", "journal_entry"])
